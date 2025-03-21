import json
import sys
import openpyxl
import requests
import os
from openpyxl import Workbook
from dotenv import load_dotenv,dotenv_values
from rich.console import Console
console = Console()
error_console = Console(stderr=True,style="bold red")

load_dotenv()

def main():
    excel_file = take_file_input_output()

    fromDate = "2024-10-13"
    toDate = "2024-10-15"

    # Convert the JSON to Excel
    json_to_excel(excel_file,fromDate,toDate)

def take_file_input_output():
    if len(sys.argv) != 2:
        error_console.log("Provide name for excel output file")
        sys.exit(1)

    output_file = sys.argv[1]

    return output_file if output_file.find(".xls") == -1 and output_file.find(".xlsx") == -1 else output_file.split(".") [0]

def login():
    url = os.getenv("LOGIN_URL")
    payload = json.dumps({
        "username": os.getenv("LOGIN_USERNAME"),
        "password": os.getenv("LOGIN_PWD"),
        "orgType": "vc"
    })
    headers = {
        'Content-Type': 'application/json'
    }
    
    response = requests.request("POST", url, headers=headers, data=payload)

    if response.headers["X-AUTH-TOKEN"] == (None or ""):
        error_console.log("Failed to login")
        sys.exit(1)
    return response.headers["X-AUTH-TOKEN"]

def fetch_user_action_data(fromDate,toDate):
    with console.status("Logging in..."):
        AUTH_TOKEN = login()
    console.log("Successfully logged in")
    
    url = "https://stringsprodapi.azure-api.net/user/api/v1/um/user_action_tracker/get_user_actions"

    payload = json.dumps({
        "userId": "",
        "orgId": "",
        "date": {
            "fromDate": fromDate,
            "toDate": toDate
        }
    })
    headers = {
        'X-AUTH-TOKEN': AUTH_TOKEN,
        'Content-Type': 'application/json'
    }

    try:
        with console.status("Fetching data"):
            response = requests.request("POST", url, headers=headers, data=payload)
        console.log("Successfully fetched data")
    except:
        error_console.log("Problem fetching data")
        sys.exit(1)
    data = response.json()

    if "response" not in data:
        error_console.log("Json does not contain response field")
        sys.exit(1)
    else:
        data = data["response"]

    return data

# Function to convert JSON data to Excel
def json_to_excel(excel_file,fromDate,toDate):
    data = fetch_user_action_data(fromDate,toDate)
    # Create a new Excel file or load an existing one
    try:
        workbook = openpyxl.load_workbook(excel_file + ".xlsx")
        worksheet = workbook.active
    except FileNotFoundError:
        workbook = Workbook()
        worksheet = workbook.active

    headers, additional_headers = add_headers_in_excel(data, worksheet)

    with console.status("Writing data to excel..."):
        write_data_to_excel(data, headers, worksheet, additional_headers)

    # Save the workbook
    workbook.save(excel_file + ".xlsx")
    console.log(f"Excel file '{excel_file}' created or updated successfully!")

def add_headers_in_excel(data, worksheet):
    headers = [
        "adminName", "date", "orgName", "adminId", "action", 
        "module", "name", "id", "type", "data", 
        "userName", "userId", "orgId"
    ]

    # Write headers only if the first row is empty
    if worksheet.max_row == 1:
        for col_num, header in enumerate(headers, start=1):
            worksheet.cell(row=1, column=col_num, value=header)

    # Collect additional headers from nested data
    additional_headers = []

    for item in data:
        for key, value in item.items():
            if isinstance(value, dict):
                for child_key in value.keys():
                    new_header = f"{key}.{child_key}"
                    if new_header not in additional_headers:
                        additional_headers.append(new_header)
            elif isinstance(value, list):
                for index, child_value in enumerate(value):
                    if isinstance(child_value, dict):
                        for child_key in child_value.keys():
                            new_header = f"{key}[{index}].{child_key}"
                            if new_header not in additional_headers:
                                additional_headers.append(new_header)

    # Write additional headers if they do not exist
    for col_num, header in enumerate(additional_headers, start=len(headers) + 1):
        if worksheet.cell(row=1, column=col_num).value is None:
            worksheet.cell(row=1, column=col_num, value=header)

    return headers, additional_headers

def write_data_to_excel(data, headers, worksheet, additional_headers):
    # Start writing data from the first empty row
    row_start = worksheet.max_row + 1

    for row_num, item in enumerate(data, start=row_start):
        for col_num, header in enumerate(headers, start=1):
            value = item.get(header)
            if isinstance(value, (dict, list)):
                worksheet.cell(row=row_num, column=col_num, value="")
            else:
                worksheet.cell(row=row_num, column=col_num, value=value)

        for col_num, header in enumerate(additional_headers, start=len(headers) + 1):
            key, child_key = header.split('.', 1) if '.' in header else (header.split('[', 1)[0], header.split('.', 1)[-1])
            if '.' in header:
                value = item.get(key, {}).get(child_key)
                if isinstance(value, (dict, list)):
                    value = json.dumps(value)
            elif '[' in header:
                index = int(header.split('[')[1].split(']')[0])
                key = header.split('[')[0]
                value = item.get(key, [])[index] if index < len(item.get(key, [])) else None
                if isinstance(value, (dict, list)):
                    value = json.dumps(value)
            else:
                value = item.get(header)

            worksheet.cell(row=row_num, column=col_num, value=value)

main()
